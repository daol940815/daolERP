#!/usr/bin/env python3
# CRM 이행 검증 — 3단 대조 (설계 §5-⑥)
#  1단: 엑셀 규칙을 파이썬으로 재현 → 통계 시트 자체 값(AU 합계·AR/AS/AT 거래플래그)과 대조
#       → 규칙을 정확히 옮겨 적었는지 증명
#  2단: 확정 규칙(§4) 파이썬 계산 → DB RPC(crm_contact_stats) 결과와 고객 단위 전수 대조
#       → SQL 구현이 정확한지 증명 (시뮬레이션 DB 필요: sim_load_orders.py --season-direct)
#  3단: 엑셀 규칙 vs 확정 규칙 분포 차이 정량화 (VIP·샘플 제외 영향)
#
# 사용: python3 validate.py <엑셀경로> [psql 접속옵션...]
#   예: python3 validate.py 고객관리.xlsx -h /tmp/pgcrm -p 5433 -U postgres -d daol
import sys, os, subprocess, io, csv
from collections import Counter, defaultdict
from openpyxl import load_workbook

EXCLUDE_ITEM = ('선결제', '퀵', '택배비', '배송비')
SEASONS = {'24설': 2024, '24추석': 2024, '25설': 2025, '25추석': 2025, '26설': 2026, '26추석': 2026}

def revenue_grade(total):
    return 'A' if total >= 10_000_000 else 'B' if total >= 5_000_000 else 'C' if total >= 1_000_000 else 'D'

def continuity_grade(b):  # b: {(year,'season'|'regular'): amount}, 기준 2026
    szn = [b.get((y, 'season'), 0) > 0 for y in (2024, 2025, 2026)]
    reg = [b.get((y, 'regular'), 0) > 0 for y in (2024, 2025, 2026)]
    if all(szn) and all(reg): return 'A'
    if szn[2] and reg[2]: return 'B'
    if all(szn) or all(reg) or szn[2] or reg[2]: return 'C'
    return 'D'

def main():
    xlsx = sys.argv[1]
    psql_opts = sys.argv[2:] or ['-h', '/tmp/pgcrm', '-p', '5433', '-U', 'postgres', '-d', 'daol']
    wb = load_workbook(xlsx, data_only=True, read_only=True)

    # ── 원장 라인 로드 (raw: SUMIFS 재현용 / stripped: 확정 규칙용) ──
    lines = []
    for r in wb['data'].iter_rows(min_row=2, values_only=True):
        if r[1] not in (2024, 2025, 2026):
            continue
        lines.append({
            'year': int(r[1]), 'month': int(r[2] or 0),
            'cancel': str(r[0] or '').strip().lower() == 'cancel',
            'bank_raw': str(r[4]) if r[4] is not None else '',
            'branch_raw': str(r[5]) if r[5] is not None else '',
            'mgr_raw': str(r[6]) if r[6] is not None else '',
            'bank': str(r[4] or '').strip(), 'branch': str(r[5] or '').strip(),
            'mgr': str(r[6] or '').strip(),
            'item': str(r[10] or '').strip(), 'kind': str(r[11] or '').strip(),
            'sale': r[14] or 0, 'pur': r[19] or 0,
            'amount': int(r[16] or 0), 'gubun': str(r[23] or '').strip(),
        })

    # ── 1단: 엑셀 규칙 재현 vs 통계 시트 캐시값 ─────────
    by_key_raw = defaultdict(list)
    for l in lines:
        by_key_raw[(l['bank_raw'], l['branch_raw'], l['mgr_raw'])].append(l)

    def excel_buckets(ls):  # 엑셀 SUMIFS: cancel·품명4종 제외, 구분=명절/상시만
        b = defaultdict(int)
        for l in ls:
            if l['cancel'] or any(x in l['item'] for x in EXCLUDE_ITEM):
                continue
            if l['gubun'] in SEASONS:
                b[(SEASONS[l['gubun']], 'season')] += l['amount']
            elif l['gubun'] == '상시':
                b[(l['year'], 'regular')] += l['amount']
        return b

    n_rows = au_mismatch = flag_mismatch = grade_mismatch = 0
    stat_keys = set()
    for r in wb['통계'].iter_rows(min_row=20, values_only=True):
        if r[0] is None:
            continue
        n_rows += 1
        key = (str(r[1]) if r[1] is not None else '',
               str(r[2]) if r[2] is not None else '',
               (str(r[3] or '') + str(r[4] or '')))
        stat_keys.add((key[0].strip(), key[1].strip(), key[2].strip()))
        b = excel_buckets(by_key_raw.get(key, []))
        total = sum(b.values())
        traded = [
            (b.get((y, 'season'), 0) + b.get((y, 'regular'), 0)) > 0 for y in (2024, 2025, 2026)
        ]
        if total != int(r[46] or 0):
            au_mismatch += 1
        if traded != [bool(int(r[43] or 0)), bool(int(r[44] or 0)), bool(int(r[45] or 0))]:
            flag_mismatch += 1
        if (revenue_grade(total), continuity_grade(b)) != (str(r[47] or ''), str(r[48] or '') or 'D'):
            grade_mismatch += 1
    print(f'[1단] 통계 {n_rows}행: 합계 불일치 {au_mismatch} / 거래플래그 불일치 {flag_mismatch} / 등급 불일치 {grade_mismatch}')

    # ── 2단: 확정 규칙 파이썬 vs DB RPC 전수 대조 ───────
    def confirmed_buckets(ls):  # 확정 규칙: + VIP·샘플 제외 (설계 §4)
        b = defaultdict(int)
        for l in ls:
            if l['cancel'] or any(x in l['item'] for x in EXCLUDE_ITEM):
                continue
            if l['item'] == 'VIP' and l['sale'] == l['pur']:
                continue
            if l['kind'] == '샘플':
                continue
            if l['year'] == 2024:
                # 2024는 legacy 경로: 구분 명절/상시만 (기타 구분 미집계 — 이행 스크립트와 동일)
                if l['gubun'] in SEASONS:
                    b[(SEASONS[l['gubun']], 'season')] += l['amount']
                elif l['gubun'] == '상시':
                    b[(l['year'], 'regular')] += l['amount']
            else:
                # 2025~: DB 경로 — 모든 라인이 erp_orders에 있음. 시즌은 구분, 나머지는 상시
                if l['gubun'] in SEASONS:
                    b[(SEASONS[l['gubun']], 'season')] += l['amount']
                else:
                    b[(l['year'], 'regular')] += l['amount']
        return b

    by_key = defaultdict(list)
    for l in lines:
        by_key[(l['bank'], l['branch'], l['mgr'])].append(l)
    py = {}
    for key, ls in by_key.items():
        b = confirmed_buckets(ls)
        # 2024 순합 0 버킷은 legacy에 저장하지 않음 → 거래여부 판정에서도 제외 (이행 스크립트와 동일)
        b = {k: v for k, v in b.items() if not (k[0] == 2024 and v == 0)}
        if b:
            py[key] = b

    q = ("COPY (SELECT k.bank_name, k.branch_name, k.manager_name, s.total_revenue,"
         " s.revenue_grade, s.continuity_grade, s.traded_y2::int, s.traded_y1::int, s.traded_y0::int"
         " FROM crm_contact_stats(2026) s JOIN crm_contact_keys k ON k.contact_id = s.contact_id)"
         " TO STDOUT WITH CSV")
    out = subprocess.run(['psql', *psql_opts, '-c', q], capture_output=True, text=True)
    if out.returncode != 0:
        sys.exit(f'DB 조회 실패: {out.stderr}')
    db_total = db_grade = db_flag = 0
    n_db = 0
    matched_keys = set()
    for row in csv.reader(io.StringIO(out.stdout)):
        bank, branch, mgr, total, rg, cg, t2, t1, t0 = row
        key = (bank, branch, mgr)
        matched_keys.add(key)
        n_db += 1
        b = py.get(key, {})
        exp_total = sum(b.values())
        exp_traded = [(b.get((y, 'season'), 0) + b.get((y, 'regular'), 0)) > 0 for y in (2024, 2025, 2026)]
        if exp_total != int(total):
            db_total += 1
        if (revenue_grade(exp_total), continuity_grade(b)) != (rg, cg):
            db_grade += 1
        if exp_traded != [t2 == '1', t1 == '1', t0 == '1']:
            db_flag += 1
    missing = {k for k in py if k not in matched_keys}
    print(f'[2단] DB 고객 {n_db}명 전수 대조: 합계 불일치 {db_total} / 등급 불일치 {db_grade} /'
          f' 거래플래그 불일치 {db_flag} / DB에 없는 파이썬 키 {len(missing)}')
    if missing:
        for k in list(missing)[:5]:
            print('   예:', k)

    # ── 3단: 엑셀 규칙 vs 확정 규칙 분포 차이 ───────────
    def distro(bucket_fn, keys):
        rev, cont = Counter(), Counter()
        for key in keys:
            b = bucket_fn(by_key.get(key, []))
            rev[revenue_grade(sum(b.values()))] += 1
            cont[continuity_grade(b)] += 1
        return rev, cont
    all_keys = set(py) | stat_keys
    er, ec = distro(excel_buckets, all_keys)
    cr, cc = distro(confirmed_buckets, all_keys)
    print(f'[3단] 동일 모집단 {len(all_keys)}명 기준')
    print(f'  매출등급  엑셀규칙 {dict(sorted(er.items()))} → 확정규칙 {dict(sorted(cr.items()))}')
    print(f'  연속성    엑셀규칙 {dict(sorted(ec.items()))} → 확정규칙 {dict(sorted(cc.items()))}')

if __name__ == '__main__':
    main()
