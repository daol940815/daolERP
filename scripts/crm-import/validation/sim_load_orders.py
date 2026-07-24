#!/usr/bin/env python3
# [검증 전용 — 운영 DB 사용 금지]
# 엑셀 원장(data 시트)의 2025~2026 라인을 로컬 테스트 DB의 erp_orders/erp_order_items로
# 시뮬레이션 적재한다. 엑셀에는 주문번호가 없으므로 (일자·은행·지점·담당자·구분) 단위로
# 가상 주문(SIM-n)을 만든다 — 라인 합계 기반인 CRM 등급 검증에는 그룹핑이 영향을 주지 않는다.
#
# 플래그는 운영 임포트(app/api/erp-orders/import)와 동일 규칙:
#   is_canceled = 상태 'cancel' / is_vip = 품명 'VIP' & 판매가=매입가 / is_prepayment = 품명 '선결제'
# season_code는 --season-direct면 구분에서 직접(라인 진실), 아니면 NULL(백필 SQL 테스트용).
#
# 사용: python3 sim_load_orders.py <엑셀경로> [--season-direct] | psql -h /tmp/pgcrm -p 5433 -U postgres -d daol
import sys, os
from openpyxl import load_workbook

SEASONS = {'24설', '24추석', '25설', '25추석', '26설', '26추석'}

def esc(v):
    if v is None or v == '':
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else os.environ.get('EXCEL')
    season_direct = '--season-direct' in sys.argv
    if not xlsx or not os.path.exists(xlsx):
        sys.exit('사용: python3 sim_load_orders.py <엑셀경로> [--season-direct]')
    wb = load_workbook(xlsx, data_only=True, read_only=True)

    orders = {}   # (date, bank, branch, mgr, gubun) → {fields, items:[]}
    for r in wb['data'].iter_rows(min_row=2, values_only=True):
        if r[1] not in (2025, 2026):
            continue
        try:
            date = f'{int(r[1])}-{int(r[2]):02d}-{int(r[3]):02d}'
        except (TypeError, ValueError):
            continue
        bank = str(r[4] or '').strip()
        branch = str(r[5] or '').strip()
        mgr = str(r[6] or '').strip()
        gubun = str(r[23] or '').strip()
        okey = (date, bank, branch, mgr, gubun)
        g = orders.setdefault(okey, {'items': []})
        item = str(r[10] or '').strip()
        sale = r[14] or 0
        pur = r[19] or 0
        g['items'].append({
            'is_canceled': str(r[0] or '').strip().lower() == 'cancel',
            'is_vip': item == 'VIP' and sale == pur,
            'is_prepayment': item == '선결제',
            'item_code': str(r[9] or '').strip() or None,
            'item_name': item or None,
            'order_kind': str(r[11] or '').strip() or None,
            'quantity': int(r[12] or 0),
            'sale_price': int(sale or 0),
            'line_total': int(r[16] or 0),
        })

    out = sys.stdout
    out.write('BEGIN;\n')
    out.write("DELETE FROM erp_order_items; DELETE FROM erp_orders;\n")  # 시뮬 DB 초기화
    n = 0
    for (date, bank, branch, mgr, gubun), g in sorted(orders.items()):
        n += 1
        season = gubun if (season_direct and gubun in SEASONS) else None
        total = sum(i['line_total'] for i in g['items'] if not i['is_canceled'])
        out.write(
            'INSERT INTO erp_orders (order_no, order_date, bank_name, branch_name, manager_name,'
            ' total_amount, collect_status, season_code)\n'
            f"VALUES ('SIM-{n:06d}', '{date}', {esc(bank)}, {esc(branch or None)}, {esc(mgr or None)},"
            f" {total}, 'collected', {esc(season)});\n")
        for ln, i in enumerate(g['items'], 1):
            out.write(
                'INSERT INTO erp_order_items (order_id, line_no, is_canceled, is_vip, is_prepayment,'
                ' item_code, item_name, order_kind, quantity, sale_price, line_total)\n'
                f"SELECT id, {ln}, {i['is_canceled']}, {i['is_vip']}, {i['is_prepayment']},"
                f" {esc(i['item_code'])}, {esc(i['item_name'])}, {esc(i['order_kind'])},"
                f" {i['quantity']}, {i['sale_price']}, {i['line_total']}"
                f" FROM erp_orders WHERE order_no = 'SIM-{n:06d}';\n")
    out.write('COMMIT;\n')
    sys.stderr.write(f'가상 주문 {n}건 생성 (2025~2026 라인 {sum(len(g["items"]) for g in orders.values())}건)\n')

if __name__ == '__main__':
    main()
