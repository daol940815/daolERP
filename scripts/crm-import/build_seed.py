#!/usr/bin/env python3
# 고객관리(CRM) 초기 이행 SQL 생성 — docs/customer-management-design.md §5
#
# 입력: 엑셀 「(주)다올커머스 고객관리」 (시트: 통계 / data / 2025 / 2026)
# 출력: out/ 아래 SQL 3개 (개인정보 포함 → out/은 git 제외)
#   01_contacts.sql        통계 시트 상세 3,011명 → crm_contacts + crm_contact_keys
#   02_legacy_2024.sql     data 시트 2024년 → crm_legacy_sales (확정 매출 인정 규칙 적용)
#   03_season_backfill.sql data 시트 구분(25설~26설) → erp_orders.season_code 백필
#
# 사용: python3 build_seed.py <엑셀경로>
# 적용: psql(로컬/테스트) 또는 Supabase SQL Editor에서 01 → 02 순서로 실행,
#       03은 erp_orders 업로드가 끝난 DB에서 실행 후 SELECT crm_match_orders();
#
# 멱등: id는 키 기반 uuid5 고정 → 재실행해도 중복 생성되지 않는다.
import sys, os, uuid, datetime, re
from openpyxl import load_workbook

NS = uuid.UUID('da01e2b0-0000-5000-8000-000000000001')  # 고정 네임스페이스 (daol CRM)
# 2024년 날짜 라인은 구분이 어느 명절이든 legacy로 (예: 2024-12 주문의 '25설' — DB 미보유 기간)
SEASONS_ALL = {'24설', '24추석', '25설', '25추석', '26설', '26추석'}
# 매출 인정 제외 품명 (엑셀 SUMIFS와 동일)
EXCLUDE_ITEM = ('선결제', '퀵', '택배비', '배송비')
TITLE_RE = re.compile(r'^(.*?)[\s]+([A-Za-z가-힣]{1,10})님?$')

def esc(v):
    if v is None or v == '':
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

def parse_name_title(raw):
    """'안아영 차장님' → ('안아영', '차장') / '김선일 님' → ('김선일', None)"""
    s = str(raw or '').strip()
    if not s:
        return None, None
    m = TITLE_RE.match(s)
    if m and m.group(1).strip():
        name, title = m.group(1).strip(), m.group(2).strip()
        if title in ('님', ''):
            title = None
        return name, title
    return s.rstrip('님').strip() or s, None

def contact_uuid(bank, branch, manager):
    return str(uuid.uuid5(NS, f'crm|{bank}|{branch}|{manager}'))

def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else os.environ.get('EXCEL')
    if not xlsx or not os.path.exists(xlsx):
        sys.exit('사용: python3 build_seed.py <고객관리 엑셀 경로>')
    outdir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'out')
    os.makedirs(outdir, exist_ok=True)
    wb = load_workbook(xlsx, data_only=True, read_only=True)

    # ── 1) 통계 시트 → contacts + keys ────────────────
    contacts = {}   # key(bank,branch,manager_raw) → dict
    for r in wb['통계'].iter_rows(min_row=20, values_only=True):
        if r[0] is None:
            continue
        bank = str(r[1] or '').strip()
        branch = str(r[2] or '').strip()
        mgr_raw = (str(r[3] or '') + str(r[4] or '')).strip()  # 지점장 or 실무자 (엑셀 키와 동일)
        if not bank:
            continue
        name, title = parse_name_title(mgr_raw)
        intimacy = str(r[49] or '').strip() or None            # AX 친밀도 (수기)
        contacts[(bank, branch, mgr_raw)] = {
            'id': contact_uuid(bank, branch, mgr_raw),
            'bank': bank, 'branch': branch, 'mgr_raw': mgr_raw,
            'name': name or '(미상)', 'title': title,
            'role': 'branch_manager' if r[3] else 'staff',
            'phone': str(r[5] or '').strip() or None,
            'counselor_prev': str(r[6] or '').strip() or None,
            'counselor_now': str(r[7] or '').strip() or None,
            'intimacy': intimacy if intimacy in ('A', 'B', 'C', 'D') else None,
            'memo': None,
        }
    n_stats = len(contacts)

    # ── 2) data 시트 2024년 → legacy 버킷 집계 ────────
    legacy = {}     # (key, season_or_None, month_or_None) → amount
    n_2024_lines = n_2024_used = 0
    unmatched_new = 0
    for r in wb['data'].iter_rows(min_row=2, values_only=True):
        if r[1] != 2024:
            continue
        n_2024_lines += 1
        status = str(r[0] or '').strip().lower()
        item = str(r[10] or '').strip()
        kind = str(r[11] or '').strip()
        gubun = str(r[23] or '').strip()
        sale = r[14] or 0
        pur = r[19] or 0
        amount = int(r[16] or 0)
        # 확정 매출 인정 규칙 (설계 §4): 취소·선결제·VIP·샘플·배송비류 제외
        if status == 'cancel':
            continue
        if any(x in item for x in EXCLUDE_ITEM):
            continue
        if item == 'VIP' and sale == pur:
            continue
        if kind == '샘플':
            continue
        if gubun in SEASONS_ALL:
            bucket = (gubun, None)
        elif gubun == '상시':
            bucket = (None, f'2024-{int(r[2]):02d}')
        else:
            continue  # 샘플/선결제/VIP/도매 등 기타 구분 — 엑셀 통계와 동일하게 미집계
        key = (str(r[4] or '').strip(), str(r[5] or '').strip(), str(r[6] or '').strip())
        if key not in contacts:  # 통계 시트에 없는 키 → 고객 자동 생성 (이관 출처 명시)
            name, title = parse_name_title(key[2])
            contacts[key] = {
                'id': contact_uuid(*key), 'bank': key[0], 'branch': key[1], 'mgr_raw': key[2],
                'name': name or '(미상)', 'title': title, 'role': 'staff',
                'phone': None, 'counselor_prev': None, 'counselor_now': None,
                'intimacy': None, 'memo': '2024 원장 이관 시 자동 생성 (통계 시트에 없던 키)',
            }
            unmatched_new += 1
        legacy[(key, *bucket)] = legacy.get((key, *bucket), 0) + amount
        n_2024_used += 1

    # ── 2-b) 2025~26 원장에만 있는 키 → 고객 자동 생성 ─
    # 통계 시트가 담당자 표기 변경('이현숙 대표님'→'이현숙 님') 등으로 빠뜨린 고객들.
    # 자동 병합은 하지 않는다(운영 원칙) — 동일인 여부는 화면에서 병합으로 확정.
    unmatched_recent = 0
    for r in wb['data'].iter_rows(min_row=2, values_only=True):
        if r[1] not in (2025, 2026):
            continue
        key = (str(r[4] or '').strip(), str(r[5] or '').strip(), str(r[6] or '').strip())
        if not key[0] or key in contacts:
            continue
        name, title = parse_name_title(key[2])
        contacts[key] = {
            'id': contact_uuid(*key), 'bank': key[0], 'branch': key[1], 'mgr_raw': key[2],
            'name': name or '(미상)', 'title': title, 'role': 'staff',
            'phone': None, 'counselor_prev': None, 'counselor_now': None,
            'intimacy': None, 'memo': '2025~26 원장 이관 시 자동 생성 (통계 시트에 없던 키)',
        }
        unmatched_recent += 1

    # ── 3) SQL 출력: contacts + keys ──────────────────
    with open(f'{outdir}/01_contacts.sql', 'w') as f:
        f.write(f'-- 생성: build_seed.py / 고객 {len(contacts)}명 (통계 {n_stats} + 자동생성 {unmatched_new + unmatched_recent})\n')
        f.write('BEGIN;\n')
        for c in contacts.values():
            f.write(
                'INSERT INTO crm_contacts (id, bank_name, branch_name, name, title, role, phone,'
                ' counselor_prev, counselor_now, intimacy_grade, memo)\n'
                f"VALUES ('{c['id']}', {esc(c['bank'])}, {esc(c['branch'] or None)}, {esc(c['name'])},"
                f" {esc(c['title'])}, '{c['role']}', {esc(c['phone'])}, {esc(c['counselor_prev'])},"
                f" {esc(c['counselor_now'])}, {esc(c['intimacy'])}, {esc(c['memo'])})\n"
                'ON CONFLICT (id) DO NOTHING;\n')
            branch_key = esc(c['branch']) if c['branch'] else "''"   # 키의 branch/manager는 NOT NULL('')
            mgr_key = esc(c['mgr_raw']) if c['mgr_raw'] else "''"
            f.write(
                'INSERT INTO crm_contact_keys (contact_id, bank_name, branch_name, manager_name, source)\n'
                f"VALUES ('{c['id']}', {esc(c['bank'])}, {branch_key}, {mgr_key}, 'import')\n"
                'ON CONFLICT (bank_name, branch_name, manager_name) DO NOTHING;\n')
        f.write('COMMIT;\n')

    # ── 4) SQL 출력: legacy 2024 ──────────────────────
    with open(f'{outdir}/02_legacy_2024.sql', 'w') as f:
        f.write(f'-- 생성: build_seed.py / 2024 라인 {n_2024_lines} 중 집계 반영 {n_2024_used}, 버킷 {len(legacy)}\n')
        f.write('BEGIN;\n')
        for (key, season, month), amount in sorted(legacy.items(), key=lambda x: str(x[0])):
            if amount == 0:
                continue  # 순합 0 버킷은 거래여부·금액 모두에 영향 없음
            cid = contacts[key]['id']
            if season:
                f.write(
                    'INSERT INTO crm_legacy_sales (contact_id, season_code, sales_month, amount)\n'
                    f"VALUES ('{cid}', {esc(season)}, NULL, {amount})\n"
                    'ON CONFLICT (contact_id, season_code) WHERE season_code IS NOT NULL'
                    ' DO UPDATE SET amount = EXCLUDED.amount;\n')
            else:
                f.write(
                    'INSERT INTO crm_legacy_sales (contact_id, season_code, sales_month, amount)\n'
                    f"VALUES ('{cid}', NULL, {esc(month)}, {amount})\n"
                    'ON CONFLICT (contact_id, sales_month) WHERE sales_month IS NOT NULL'
                    ' DO UPDATE SET amount = EXCLUDED.amount;\n')
        f.write('COMMIT;\n')

    # ── 5) SQL 출력: 시즌 백필 (erp_orders) ───────────
    # 엑셀 라인의 구분을 (주문일, 은행, 지점, 담당자) 단위로 모아 명절을 유도.
    # 같은 키·날짜에 서로 다른 명절이 섞이면 건너뛴다(수동 확인 대상으로 보고).
    smap = {}       # (date, bank, branch, mgr) → set(seasons)
    for sheet in ('data',):
        for r in wb[sheet].iter_rows(min_row=2, values_only=True):
            if r[1] not in (2025, 2026):
                continue
            gubun = str(r[23] or '').strip()
            if gubun not in SEASONS_ALL:
                continue
            try:
                d = datetime.date(int(r[1]), int(r[2]), int(r[3]))
            except (TypeError, ValueError):
                continue
            key = (d, str(r[4] or '').strip(), str(r[5] or '').strip(), str(r[6] or '').strip())
            smap.setdefault(key, set()).add(gubun)
    rows = [(k, next(iter(v))) for k, v in smap.items() if len(v) == 1]
    ambiguous = [k for k, v in smap.items() if len(v) > 1]
    with open(f'{outdir}/03_season_backfill.sql', 'w') as f:
        f.write(f'-- 생성: build_seed.py / 백필 대상 키 {len(rows)}, 다중 명절 혼재로 제외 {len(ambiguous)}\n')
        for k in ambiguous:
            f.write(f'-- 제외(혼재): {k[0]} {k[1]} {k[2]} {k[3]}\n')
        f.write('BEGIN;\n')
        f.write('CREATE TEMP TABLE _crm_season_map ('
                'order_date DATE, bank TEXT, branch TEXT, manager TEXT, season VARCHAR(10)) ON COMMIT DROP;\n')
        CHUNK = 500
        def esc_nn(v):  # 맵의 bank/branch/manager는 빈 문자열 유지 (COALESCE(o.컬럼,'') 조인 대상)
            return "'" + str(v).replace("'", "''") + "'"
        vals = [f"('{k[0]}', {esc_nn(k[1])}, {esc_nn(k[2])}, {esc_nn(k[3])}, {esc(s)})" for k, s in rows]
        for i in range(0, len(vals), CHUNK):
            f.write('INSERT INTO _crm_season_map VALUES\n' + ',\n'.join(vals[i:i+CHUNK]) + ';\n')
        f.write(
            'UPDATE erp_orders o SET season_code = m.season\n'
            'FROM _crm_season_map m\n'
            'WHERE o.order_date = m.order_date\n'
            "  AND COALESCE(o.bank_name, '')    = m.bank\n"
            "  AND COALESCE(o.branch_name, '')  = m.branch\n"
            "  AND COALESCE(o.manager_name, '') = m.manager\n"
            '  AND o.season_code IS DISTINCT FROM m.season;\n')
        f.write('COMMIT;\n')

    print(f'고객 {len(contacts)}명 (통계 {n_stats} + 2024 자동생성 {unmatched_new})')
    print(f'2024 legacy: 라인 {n_2024_lines} → 반영 {n_2024_used}, 버킷 {sum(1 for a in legacy.values() if a != 0)}')
    print(f'시즌 백필: 키 {len(rows)}, 혼재 제외 {len(ambiguous)}')
    print(f'출력: {outdir}/01_contacts.sql, 02_legacy_2024.sql, 03_season_backfill.sql')

if __name__ == '__main__':
    main()
